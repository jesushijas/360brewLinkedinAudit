#!/usr/bin/env python3
"""
360Brew Scorecard Generator

Usage:
    python generate_scorecard.py --data '<JSON>' --output /path/to/output.xlsx --user "Nombre Usuario" --handle "@handle"

The --data argument expects a JSON string with this structure:
[
  {
    "category": "1. HEADLINE (Titular)",
    "items": [
      {
        "criterion": "Densidad semántica",
        "what": "Descripción de qué evalúa...",
        "weight": 5,
        "score": 3,
        "diagnosis": "Diagnóstico específico del perfil..."
      }
    ]
  }
]

Each item requires: criterion (str), what (str), weight (int 2-10), score (int 1-5), diagnosis (str).
Categories with no items are skipped.
"""

import json
import sys
import argparse
from datetime import datetime

def main():
    parser = argparse.ArgumentParser(description='Generate 360Brew Scorecard Excel')
    parser.add_argument('--data', required=True, help='JSON string with scoring data')
    parser.add_argument('--output', required=True, help='Output .xlsx path')
    parser.add_argument('--user', default='Usuario', help='User display name')
    parser.add_argument('--handle', default='', help='LinkedIn handle (e.g. @username)')
    args = parser.parse_args()

    try:
        categories = json.loads(args.data)
    except json.JSONDecodeError as e:
        print(f"Error parsing JSON: {e}", file=sys.stderr)
        sys.exit(1)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("Installing openpyxl...", file=sys.stderr)
        import subprocess
        subprocess.check_call([sys.executable, '-m', 'pip', 'install',
                               'openpyxl', '--break-system-packages', '-q'])
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "360Brew Scorecard"

    # ── Colors ──
    def h(hex_str): return hex_str.lstrip('#')

    BG_DARK = h('#0A0E1A')
    BG_HEADER = h('#0A66C2')
    BG_HEADER2 = h('#004182')
    BG_CAT = h('#1A2238')
    BG_ROW1 = h('#111827')
    BG_ROW2 = h('#0F1525')
    BG_SCORE_GREEN = h('#064E3B')
    BG_SCORE_YELLOW = h('#78350F')
    BG_SCORE_RED = h('#7F1D1D')
    BG_TOTAL = h('#1E3A5F')
    WHITE = h('#FFFFFF')
    GREY = h('#9CA3AF')
    BLUE_LIGHT = h('#93C5FD')
    GREEN = h('#34D399')
    YELLOW = h('#FBBF24')
    RED = h('#F87171')
    CYAN = h('#67E8F9')

    thin = Border(
        left=Side(style='thin', color='1F2937'),
        right=Side(style='thin', color='1F2937'),
        top=Side(style='thin', color='1F2937'),
        bottom=Side(style='thin', color='1F2937')
    )

    def cell(ws, r, c, val, bold=False, size=10, color=WHITE, bg=BG_DARK,
             halign='left', valign='center', wrap=True):
        cl = ws.cell(row=r, column=c, value=val)
        cl.font = Font(name='Arial', bold=bold, size=size, color=color)
        cl.fill = PatternFill('solid', start_color=bg)
        cl.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
        cl.border = thin
        return cl

    def score_bg(score, max_score):
        pct = score / max_score if max_score > 0 else 0
        if pct >= 0.7: return BG_SCORE_GREEN
        if pct >= 0.4: return BG_SCORE_YELLOW
        return BG_SCORE_RED

    def score_color(score, max_score):
        pct = score / max_score if max_score > 0 else 0
        if pct >= 0.7: return GREEN
        if pct >= 0.4: return YELLOW
        return RED

    # Column widths
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 52
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 52

    # ── TITLE ──
    ws.merge_cells('A1:F1')
    cell(ws, 1, 1, '360BREW SCORECARD — EVALUACIÓN DE PERFIL LINKEDIN',
         bold=True, size=14, color=WHITE, bg=BG_HEADER, halign='center')
    ws.row_dimensions[1].height = 32

    date_str = datetime.now().strftime('%B %Y')
    handle_str = f' ({args.handle})' if args.handle else ''
    ws.merge_cells('A2:F2')
    cell(ws, 2, 1,
         f'Perfil evaluado: {args.user}{handle_str} · Fecha: {date_str} · Evaluador: Claude',
         size=9, color=GREY, bg=BG_DARK, halign='center')
    ws.row_dimensions[2].height = 20

    # ── HEADERS ──
    headers = ['#', 'CRITERIO', 'QUÉ EVALÚA Y POR QUÉ IMPORTA PARA 360BREW',
               'PESO\n(2-10)', 'NOTA\n(1-5)', 'DIAGNÓSTICO']
    for i, hdr in enumerate(headers, 1):
        cell(ws, 3, i, hdr, bold=True, size=9, color=WHITE, bg=BG_HEADER2, halign='center')
    ws.row_dimensions[3].height = 30

    # ── DATA ──
    row = 4
    criteria_num = 0

    for cat in categories:
        if not cat.get('items'):
            continue

        ws.merge_cells(f'A{row}:F{row}')
        cell(ws, row, 1, cat['category'], bold=True, size=10, color=CYAN,
             bg=BG_CAT, halign='left')
        ws.row_dimensions[row].height = 24
        row += 1

        for item in cat['items']:
            criteria_num += 1
            bg = BG_ROW1 if criteria_num % 2 == 0 else BG_ROW2
            row_h = max(60, len(item.get('what', '')) // 3 + 20)
            ws.row_dimensions[row].height = row_h

            cell(ws, row, 1, criteria_num, size=9, color=GREY, bg=bg, halign='center')
            cell(ws, row, 2, item['criterion'], bold=True, size=10, color=WHITE, bg=bg)
            cell(ws, row, 3, item.get('what', ''), size=9, color=GREY, bg=bg)
            cell(ws, row, 4, item['weight'], bold=True, size=11, color=BLUE_LIGHT,
                 bg=bg, halign='center')
            sc = item['score']
            cell(ws, row, 5, sc, bold=True, size=14,
                 color=score_color(sc, 5), bg=score_bg(sc, 5), halign='center')
            cell(ws, row, 6, item.get('diagnosis', ''), size=9, color=WHITE, bg=bg)
            row += 1

    # ── TOTALS ──
    row += 1
    ws.merge_cells(f'A{row}:C{row}')
    cell(ws, row, 1, 'PUNTUACIÓN TOTAL', bold=True, size=12, color=WHITE,
         bg=BG_TOTAL, halign='right')
    cell(ws, row, 4, '', bg=BG_TOTAL)

    total_weighted = 0
    total_weight = 0
    for cat in categories:
        for item in cat.get('items', []):
            total_weighted += item['score'] * item['weight']
            total_weight += item['weight']

    max_possible = total_weight * 5
    final_score = round((total_weighted / max_possible) * 100, 1) if max_possible > 0 else 0

    cell(ws, row, 5, final_score, bold=True, size=16,
         color=score_color(final_score, 100), bg=BG_TOTAL, halign='center')
    cell(ws, row, 6,
         f'sobre 100  (puntuación ponderada: {total_weighted} de {max_possible} puntos posibles)',
         bold=True, size=10, color=GREY, bg=BG_TOTAL)
    ws.row_dimensions[row].height = 32

    # ── LEGEND ──
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    cell(ws, row, 1, 'SISTEMA DE PUNTUACIÓN', bold=True, size=10, color=CYAN,
         bg=BG_CAT, halign='left')
    ws.row_dimensions[row].height = 22
    row += 1

    legend_items = [
        ('Nota (columna E)',
         'Cada criterio se puntúa de 1 a 5. 1 = gap crítico, 2 = debajo de lo esperado, '
         '3 = aceptable con margen de mejora, 4 = bien, 5 = óptimo.'),
        ('Peso (columna D)',
         'Refleja el impacto relativo de cada criterio en cómo 360Brew construye tu authority '
         'score. Los criterios de contenido pesan más porque 360Brew prioriza el match semántico '
         'entre lo que dices ser y lo que demuestras publicando.'),
        ('Puntuación final',
         'Se calcula como: Σ(nota × peso) / Σ(peso × 5) × 100. Es un porcentaje ponderado '
         'donde los criterios de mayor impacto en 360Brew tienen mayor peso en la nota final.'),
        ('Colores',
         'Verde (≥70%): alineado con 360Brew. Amarillo (40-69%): funcional pero con margen '
         'significativo de mejora. Rojo (<40%): gap que puede limitar tu alcance.'),
    ]

    for label, desc in legend_items:
        ws.merge_cells(f'C{row}:F{row}')
        cell(ws, row, 1, '', bg=BG_ROW1)
        cell(ws, row, 2, label, bold=True, size=9, color=BLUE_LIGHT, bg=BG_ROW1)
        cell(ws, row, 3, desc, size=9, color=GREY, bg=BG_ROW1)
        ws.row_dimensions[row].height = max(40, len(desc) // 4 + 15)
        row += 1

    # ── BREAKDOWN PER CATEGORY ──
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    cell(ws, row, 1, 'DESGLOSE POR CATEGORÍA', bold=True, size=10, color=CYAN,
         bg=BG_CAT, halign='left')
    ws.row_dimensions[row].height = 22
    row += 1

    for cat in categories:
        items = cat.get('items', [])
        if not items:
            continue
        cat_weighted = sum(i['score'] * i['weight'] for i in items)
        cat_max = sum(i['weight'] for i in items) * 5
        cat_pct = round((cat_weighted / cat_max) * 100, 1) if cat_max > 0 else 0

        cell(ws, row, 1, '', bg=BG_ROW1)
        cell(ws, row, 2, cat['category'], bold=True, size=9, color=WHITE, bg=BG_ROW1)
        ws.merge_cells(f'C{row}:D{row}')
        bar_filled = int(cat_pct / 5)
        bar = '\u2588' * bar_filled + '\u2591' * (20 - bar_filled)
        cell(ws, row, 3, bar, size=9, color=score_color(cat_pct, 100), bg=BG_ROW1)
        cell(ws, row, 5, cat_pct, bold=True, size=11,
             color=score_color(cat_pct, 100), bg=score_bg(cat_pct, 100), halign='center')
        cell(ws, row, 6, f'{cat_weighted}/{cat_max} pts', size=9, color=GREY, bg=BG_ROW1)
        ws.row_dimensions[row].height = 22
        row += 1

    ws.freeze_panes = 'A4'

    wb.save(args.output)
    print(f"Score: {final_score}/100")
    print(f"Weighted: {total_weighted}/{max_possible}")
    print(f"Saved: {args.output}")


if __name__ == '__main__':
    main()
